import os
import json
import re
import csv
import time
import secrets
import smtplib
import urllib.parse
import urllib.request
from functools import wraps
from datetime import datetime, timedelta
from io import BytesIO, StringIO
from email.message import EmailMessage

from flask import Flask, render_template as flask_render_template, request, redirect, session, send_file, Response, flash
from flask_session import Session
import pyodbc
import random

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.exceptions import HTTPException

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY")
# If env not set, keep old behavior to avoid breaking dev unexpectedly,
# but strongly prefer setting FLASK_SECRET_KEY in production.
if not app.secret_key:
    app.secret_key = "change-this-secret-key"

# ================= SESSION FIX (IMPORTANT) =================
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
app.config["SESSION_FILE_DIR"] = os.path.join(app.root_path, "flask_session")
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=30)
os.makedirs(app.config["SESSION_FILE_DIR"], exist_ok=True)

# extra hardening (doesn't change app behavior)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = False

Session(app)
SCHEMA_READY = False


TEMPLATE_ALIASES = {
    "login.html": "Login.html",
    "add_student.html": "Add_student.html",
    "students.html": "Students.html",
    "settings.html": "Settings.html",
    "error.html": "Error.html",
}


def render_template(template_name, **context):
    template_name = TEMPLATE_ALIASES.get(template_name, template_name)
    context.setdefault("profile_pic", "Default.png")
    context.setdefault(
        "stats",
        {
            "students": 0,
            "users": 0,
            "queries": 0,
            "reports": 0,
            "last_login": None,
            "recent_logs": [],
        },
    )
    context.setdefault("tables", [])
    return flask_render_template(template_name, **context)


def load_sms_config():
    config_path = os.path.join(app.root_path, "sms_config.json")
    if not os.path.exists(config_path):
        return {}

    try:
        with open(config_path, encoding="utf-8") as config_file:
            return json.load(config_file)
    except Exception as e:
        print(f"SMS config load failed: {e}")
        return {}


def load_email_config():
    config_path = os.path.join(app.root_path, "email_config.json")
    if not os.path.exists(config_path):
        return {}

    try:
        with open(config_path, encoding="utf-8") as config_file:
            return json.load(config_file)
    except Exception as e:
        print(f"Email config load failed: {e}")
        return {}


SQL_DRIVER = os.environ.get("SQL_DRIVER", "ODBC Driver 17 for SQL Server")
APP_DB_CONFIG = {
    "server": os.environ.get("APP_DB_SERVER", r"SHUBHAM\SQLEXPRESS"),
    "database": os.environ.get("APP_DB_NAME", "Shubham"),
    "uid": os.environ.get("APP_DB_UID", "sa"),
    "pwd": os.environ.get("APP_DB_PWD", "Shubham@321"),
}
DEFAULT_SERVERS = [
    APP_DB_CONFIG["server"],
   
    r".\SQLEXPRESS",
]
SMS_CONFIG = load_sms_config()
EMAIL_CONFIG = load_email_config()
SMS_PROVIDER = (
    os.environ.get("SMS_PROVIDER")
    or SMS_CONFIG.get("provider")
    or "fast2sms"
).lower()
FAST2SMS_API_KEY = (
    os.environ.get("FAST2SMS_API_KEY")
    or SMS_CONFIG.get("fast2sms_api_key")
    or ""
).strip()
SMTP_HOST = (os.environ.get("SMTP_HOST") or EMAIL_CONFIG.get("host") or "").strip()
SMTP_PORT = int(os.environ.get("SMTP_PORT") or EMAIL_CONFIG.get("port") or "587")
SMTP_USERNAME = (os.environ.get("SMTP_USERNAME") or EMAIL_CONFIG.get("username") or "").strip()
SMTP_PASSWORD = (os.environ.get("SMTP_PASSWORD") or EMAIL_CONFIG.get("password") or "").strip()
SMTP_FROM = (os.environ.get("SMTP_FROM") or EMAIL_CONFIG.get("from") or SMTP_USERNAME).strip()
SMTP_USE_TLS = str(os.environ.get("SMTP_USE_TLS", EMAIL_CONFIG.get("use_tls", "1"))).strip().lower() not in ("0", "false", "no")


def build_conn_str(server, database=None, uid=None, pwd=None):
    parts = [
        f"DRIVER={{{SQL_DRIVER}}}",
        f"SERVER={server}",
        "TrustServerCertificate=yes",
    ]
    if database:
        parts.append(f"DATABASE={database}")
    if uid:
        parts.append(f"UID={uid}")
    if pwd:
        parts.append(f"PWD={pwd}")
    return ";".join(parts) + ";"


def connect_sql_server(server, database=None, uid=None, pwd=None):
    return pyodbc.connect(build_conn_str(server, database, uid, pwd))


def get_app_conn():
    return connect_sql_server(**APP_DB_CONFIG)


def ensure_app_schema():
    global SCHEMA_READY
    if SCHEMA_READY:
        return

    conn = None
    try:
        conn = connect_sql_server(**APP_DB_CONFIG)
        cursor = conn.cursor()
        statements = [
            """
            IF COL_LENGTH('dbo.users', 'password_hash') IS NULL
                ALTER TABLE dbo.users ADD password_hash NVARCHAR(255) NULL
            """,
            """
            IF COL_LENGTH('dbo.users', 'role') IS NULL
                ALTER TABLE dbo.users ADD role NVARCHAR(30) NOT NULL CONSTRAINT DF_users_role DEFAULT 'user'
            """,
            """
            IF COL_LENGTH('dbo.users', 'is_active') IS NULL
                ALTER TABLE dbo.users ADD is_active BIT NOT NULL CONSTRAINT DF_users_is_active DEFAULT 1
            """,
            """
            IF COL_LENGTH('dbo.users', 'last_login') IS NULL
                ALTER TABLE dbo.users ADD last_login DATETIME NULL
            """,
            """
            IF COL_LENGTH('dbo.users', 'failed_attempts') IS NULL
                ALTER TABLE dbo.users ADD failed_attempts INT NOT NULL CONSTRAINT DF_users_failed_attempts DEFAULT 0
            """,
            """
            IF OBJECT_ID('dbo.activity_logs', 'U') IS NULL
            CREATE TABLE dbo.activity_logs (
                id INT IDENTITY(1,1) PRIMARY KEY,
                username NVARCHAR(100) NULL,
                action NVARCHAR(100) NOT NULL,
                details NVARCHAR(MAX) NULL,
                ip_address NVARCHAR(60) NULL,
                created_at DATETIME NOT NULL DEFAULT GETDATE()
            )
            """,
            """
            IF OBJECT_ID('dbo.saved_queries', 'U') IS NULL
            CREATE TABLE dbo.saved_queries (
                id INT IDENTITY(1,1) PRIMARY KEY,
                username NVARCHAR(100) NOT NULL,
                title NVARCHAR(150) NOT NULL,
                query_text NVARCHAR(MAX) NOT NULL,
                created_at DATETIME NOT NULL DEFAULT GETDATE()
            )
            """,
            """
            IF OBJECT_ID('dbo.report_history', 'U') IS NULL
            CREATE TABLE dbo.report_history (
                id INT IDENTITY(1,1) PRIMARY KEY,
                username NVARCHAR(100) NULL,
                report_name NVARCHAR(180) NOT NULL,
                rows_count INT NOT NULL DEFAULT 0,
                sheets_count INT NOT NULL DEFAULT 0,
                created_at DATETIME NOT NULL DEFAULT GETDATE()
            )
            """,
            """
            IF COL_LENGTH('dbo.users', 'role') IS NOT NULL
               AND NOT EXISTS (SELECT 1 FROM dbo.users WHERE role='admin')
            BEGIN
                UPDATE dbo.users
                SET role='admin', is_active=1
                WHERE username = (SELECT TOP 1 username FROM dbo.users ORDER BY username)
            END
            """,
            """
            IF COL_LENGTH('dbo.users', 'role') IS NOT NULL
               AND COL_LENGTH('dbo.users', 'is_active') IS NOT NULL
            BEGIN
                UPDATE dbo.users
                SET is_active=1
                WHERE role='admin'
            END
            """,
        ]

        for statement in statements:
            cursor.execute(statement)

        conn.commit()
        SCHEMA_READY = True
    except Exception as e:
        print(f"Schema setup skipped: {e}")
    finally:
        if conn:
            conn.close()


def normalize_mobile(mobile):
    digits = re.sub(r"\D", "", mobile or "")
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    if len(digits) == 10 and digits[0] in "6789":
        return digits
    return ""


def mask_mobile(mobile):
    digits = normalize_mobile(mobile)
    if not digits:
        return ""
    return f"******{digits[-4:]}"


def mask_email(email):
    email = (email or "").strip()
    if "@" not in email:
        return ""

    name, domain = email.split("@", 1)
    if not name or not domain:
        return ""

    if len(name) <= 2:
        masked_name = name[0] + "*"
    else:
        masked_name = name[0] + "*" * (len(name) - 2) + name[-1]

    return f"{masked_name}@{domain}"


def send_otp_email(email, otp):
    email = (email or "").strip()
    if "@" not in email:
        return False, "Registered email address not found"

    if not (SMTP_HOST and SMTP_FROM):
        print(f"OTP for {email}: {otp}")
        return True, "Email is not configured, so OTP printed in console"

    message = EmailMessage()
    message["Subject"] = "Shubham's Application"
    message["From"] = SMTP_FROM
    message["To"] = email
    message.set_content(
        f"Application login OTP is {otp}.\n\n"
        "This code is valid for 5 minutes. If you did not request it, please ignore this email."
    )

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as smtp:
            if SMTP_USE_TLS:
                smtp.starttls()
            if SMTP_USERNAME and SMTP_PASSWORD:
                smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
            smtp.send_message(message)
        return True, "OTP sent to your registered email"
    except Exception as e:
        return False, f"Email OTP failed: {e}"


def send_otp_sms(mobile, otp):
    mobile = normalize_mobile(mobile)
    if not mobile:
        return False, "Valid mobile number not found"

    if SMS_PROVIDER == "fast2sms" and FAST2SMS_API_KEY:
        query = urllib.parse.urlencode({
            "authorization": FAST2SMS_API_KEY,
            "route": "otp",
            "variables_values": otp,
            "flash": "0",
            "numbers": mobile,
        })
        url = f"https://www.fast2sms.com/dev/bulkV2?{query}"

        try:
            with urllib.request.urlopen(url, timeout=15) as response:
                response_data = json.loads(response.read().decode("utf-8"))
            if response_data.get("return") is True:
                return True, "OTP sent successfully"
            return False, response_data.get("message", "OTP SMS failed")
        except urllib.error.HTTPError as e:
            error_body = e.read().decode("utf-8", errors="replace")
            try:
                error_data = json.loads(error_body)
                error_message = error_data.get("message") or error_body
            except json.JSONDecodeError:
                error_message = error_body or str(e)
            return False, f"OTP SMS failed: {error_message}"
        except Exception as e:
            return False, f"OTP SMS failed: {e}"

    print(f"OTP for {mobile}: {otp}")
    return True, "OTP printed in console"


def login_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect("/login")
        return view_func(*args, **kwargs)

    return wrapper


def csrf_token():
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_csrf_token"] = token
    return token


@app.context_processor
def inject_global_template_helpers():
    return {"csrf_token": csrf_token}


@app.before_request
def prepare_request():
    if not request.path.startswith("/static/"):
        ensure_app_schema()


@app.before_request
def protect_post_requests():
    if request.method not in ("POST", "PUT", "PATCH", "DELETE"):
        return None

    expected = session.get("_csrf_token")
    received = request.form.get("_csrf_token") or request.headers.get("X-CSRFToken")

    if not expected or not received or not secrets.compare_digest(expected, received):
        if wants_json_response():
            payload = {"error": "Security token expired. Please refresh and try again.", "status_code": 400}
            return Response(json.dumps(payload), status=400, mimetype="application/json")
        return render_error_page(
            title="Security check failed",
            message="Security token expired. Please refresh the page and try again.",
            status_code=400,
        )


def admin_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect("/login")
        if session.get("role") != "admin":
            return render_template(
                "Error.html",
                title="Access denied",
                message="Only admin users can open this page."
            )
        return view_func(*args, **kwargs)

    return wrapper


def wants_json_response():
    json_paths = (
        "/get_databases",
        "/get_tables",
        "/save_query",
        "/delete_saved_query",
        "/admin/users/update",
    )
    if request.path.startswith(json_paths):
        return True
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return True
    return (
        request.accept_mimetypes["application/json"]
        > request.accept_mimetypes["text/html"]
    )


def render_error_page(title="Error", message="Something went wrong.", status_code=500):
    try:
        return render_template(
            "Error.html",
            title=title,
            message=message,
            status_code=status_code,
            path=request.path,
        ), status_code
    except Exception:
        return f"{title}: {message}", status_code


@app.errorhandler(HTTPException)
def handle_http_exception(error):
    status_code = error.code or 500
    title = error.name or "Error"
    message = error.description or "The requested action could not be completed."

    if wants_json_response():
        payload = {"error": message, "status_code": status_code}
        return Response(json.dumps(payload), status=status_code, mimetype="application/json")

    return render_error_page(title=title, message=message, status_code=status_code)


@app.errorhandler(Exception)
def handle_unexpected_exception(error):
    app.logger.exception("Unhandled application error")

    if wants_json_response():
        payload = {
            "error": "An unexpected error occurred. Please try again.",
            "status_code": 500,
        }
        return Response(json.dumps(payload), status=500, mimetype="application/json")

    return render_error_page(
        title="Application Error",
        message="An unexpected error occurred. Please try again or go back to the dashboard.",
        status_code=500,
    )


def get_client_ip():
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.remote_addr or ""


def log_activity(action, details=""):
    conn = None
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO dbo.activity_logs (username, action, details, ip_address)
            VALUES (?, ?, ?, ?)
            """,
            (session.get("username") or session.get("temp_user"), action, details, get_client_ip())
        )
        conn.commit()
    except Exception as e:
        print(f"Activity log failed: {e}")
    finally:
        if conn:
            conn.close()


def verify_user_password(cursor, username, plain_password):
    cursor.execute(
        """
        SELECT username, email, mobile, password, password_hash, is_active, role, failed_attempts
        FROM dbo.users
        WHERE username=?
        """,
        (username,)
    )
    user = cursor.fetchone()
    if not user:
        return None

    if user[5] is not None and int(user[5]) == 0:
        return "inactive"

    if user[7] is not None and int(user[7]) >= 5:
        return "locked"

    password_hash = user[4] or ""
    old_password = user[3] or ""
    valid = False

    if password_hash:
        try:
            valid = check_password_hash(password_hash, plain_password)
        except Exception:
            valid = False
    elif old_password == plain_password:
        valid = True
        cursor.execute(
            "UPDATE dbo.users SET password_hash=? WHERE username=?",
            (generate_password_hash(plain_password), username)
        )

    if valid:
        cursor.execute(
            "UPDATE dbo.users SET failed_attempts=0 WHERE username=?",
            (username,)
        )
        return user

    cursor.execute(
        "UPDATE dbo.users SET failed_attempts=ISNULL(failed_attempts, 0) + 1 WHERE username=?",
        (username,)
    )
    return None


def require_safe_sql(query_text):
    query = (query_text or "").strip()
    backend_proc, _ = get_report_table_name(query)
    if backend_proc:
        return

    dangerous = re.search(
        r"\b(DROP|TRUNCATE|ALTER|CREATE|MERGE|DELETE|UPDATE|INSERT)\b",
        query,
        re.IGNORECASE
    )
    if dangerous:
        raise ValueError(
            f"Dangerous SQL command blocked: {dangerous.group(1).upper()}. "
            "Use SELECT/EXEC reports only from this runner."
        )


def fetch_dashboard_stats():
    stats = {
        "students": 0,
        "users": 0,
        "queries": 0,
        "reports": 0,
        "last_login": "",
        "recent_logs": [],
    }
    conn = None
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM dbo.students")
        stats["students"] = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM dbo.users")
        stats["users"] = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM dbo.saved_queries")
        stats["queries"] = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM dbo.report_history")
        stats["reports"] = cursor.fetchone()[0]
        cursor.execute(
            "SELECT last_login FROM dbo.users WHERE username=?",
            (session.get("username"),)
        )
        row = cursor.fetchone()
        if row and row[0]:
            stats["last_login"] = row[0].strftime("%d %b %Y %I:%M %p")
        cursor.execute("""
            SELECT TOP 6 action, details, created_at
            FROM dbo.activity_logs
            ORDER BY id DESC
        """)
        stats["recent_logs"] = cursor.fetchall()
    except Exception as e:
        print(f"Dashboard stats failed: {e}")
    finally:
        if conn:
            conn.close()
    return stats


def fetch_saved_queries():
    conn = None
    queries = []
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT id, title, query_text, created_at
            FROM dbo.saved_queries
            WHERE username=?
            ORDER BY id DESC
            """,
            (session.get("username"),)
        )
        queries = cursor.fetchall()
    except Exception as e:
        print(f"Saved query load failed: {e}")
    finally:
        if conn:
            conn.close()
    return queries


def record_report_history(report_name, rows_count=0, sheets_count=0):
    conn = None
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO dbo.report_history (username, report_name, rows_count, sheets_count)
            VALUES (?, ?, ?, ?)
            """,
            (session.get("username"), report_name, rows_count, sheets_count)
        )
        conn.commit()
    except Exception as e:
        print(f"Report history failed: {e}")
    finally:
        if conn:
            conn.close()


def is_sidbi_top_sheet(item):
    columns = item.get("columns", [])
    query = item.get("query", "").lower()
    return (
        len(columns) >= 11
        and columns[0] == "Bucket"
        and columns[5].strip() == ""
        and columns[6] == "Bucket"
    ) or "usp_sidbi_prayaas_topsheet" in query


def clean_excel_sheet_name(name):
    invalid_chars = '\\/*?:[]'
    clean_name = ''.join('_' if char in invalid_chars else char for char in name)
    clean_name = clean_name.strip().strip("'") or "Sheet"
    return clean_name[:31]


def unique_excel_sheet_name(name, used_names):
    base_name = clean_excel_sheet_name(name)
    sheet_name = base_name
    counter = 2

    while sheet_name.lower() in used_names:
        suffix = f"_{counter}"
        sheet_name = f"{base_name[:31 - len(suffix)]}{suffix}"
        counter += 1

    used_names.add(sheet_name.lower())
    return sheet_name


def get_report_sheet_name(query, index):
    query_lower = query.lower()

    if "usp_sidbi_prayaas_topsheet" in query_lower:
        return "Top Sheet"
    if "usp_dpd_report_1_to_30" in query_lower:
        return "Staff Wise 1 to 30"
    if "usp_dpd_report" in query_lower:
        return "Staff Wise"

    return f"Query_{index}"


def apply_sidbi_top_sheet_format(sheet, item):
    title_font = Font(bold=True, size=11)
    header_font = Font(bold=True)
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    sheet.merge_cells("A1:E1")
    sheet.merge_cells("G1:K1")
    sheet.merge_cells("A2:E2")
    sheet.merge_cells("G2:K2")

    report_month = item.get("report_month") or "Apr-26"
    sheet["A1"] = f"Top Sheet {report_month}"
    sheet["G1"] = f"Top Sheet {report_month}"
    sheet["A2"] = "1 To 28 Days Wise"
    sheet["G2"] = "1 To 30 Days Wise"

    for row in (1, 2):
        for col in range(1, 12):
            cell = sheet.cell(row=row, column=col)
            cell.font = title_font
            cell.alignment = center
            if col != 6:
                cell.border = border

    for col_index, column_name in enumerate(item["columns"], start=1):
        cell = sheet.cell(row=3, column=col_index, value=column_name)
        cell.font = header_font
        cell.alignment = center
        if col_index != 6:
            cell.fill = yellow_fill
            cell.border = border

    for row_index, row_data in enumerate(item["rows"], start=4):
        is_total = str(row_data[0]).strip().lower() == "grand total"
        for col_index, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = center
            if col_index != 6:
                cell.border = border
            if is_total and col_index != 6:
                cell.font = header_font
                cell.fill = yellow_fill

    widths = {
        "A": 22, "B": 12, "C": 12, "D": 14, "E": 18,
        "F": 8,
        "G": 22, "H": 12, "I": 12, "J": 14, "K": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def is_staff_wise_sidbi_prayas(item):
    query = item.get("query", "").lower()
    columns = item.get("columns", [])
    return (
        "usp_dpd_report" in query
        and "usp_dpd_report_1_to_30" not in query
    ) or (
        "Staff ID" in columns
        and any(col == "ALC 0 DPD" for col in columns)
        and any(col == "Outstanding >90 DPD" or col == "Outstanding >90 DPD S3" for col in columns)
    )


def is_staff_wise_sidbi_prayas_1_to_30(item):
    query = item.get("query", "").lower()
    return "usp_dpd_report_1_to_30" in query


def apply_staff_wise_sidbi_format(sheet, item, title="STAFF WISE SIDBI PRAYAS"):
    columns = item["columns"]
    rows = item["rows"]
    max_col = len(columns)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    title_fill = PatternFill("solid", fgColor="FFC7CE")
    branch_fill = PatternFill("solid", fgColor="FFFF00")
    total_fill = PatternFill("solid", fgColor="D9EAD3")
    alc_fill = PatternFill("solid", fgColor="FFF2CC")
    loan_fill = PatternFill("solid", fgColor="D9EAF7")
    arrear_fill = PatternFill("solid", fgColor="EADCF8")
    outstanding_fill = PatternFill("solid", fgColor="DDEBF7")
    percent_fill = PatternFill("solid", fgColor="E2F0D9")
    grand_total_fill = PatternFill("solid", fgColor="FFFF00")

    title_font = Font(bold=True, size=12, color="C00000")
    group_font = Font(bold=True, size=10)
    header_font = Font(bold=True, size=9)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center

    groups = [
        ("Branch, Region And Staff Detail", 1, 5, branch_fill),
        ("Total", 6, 8, total_fill),
        ("ALC", 9, 15, alc_fill),
        ("Loan", 16, 22, loan_fill),
        ("Arrear Amount", 23, 29, arrear_fill),
        ("Outstanding", 30, 36, outstanding_fill),
        ("Outstanding %", 37, max_col, percent_fill),
    ]

    for group_name, start_col, end_col, fill in groups:
        if start_col > max_col:
            continue
        end_col = min(end_col, max_col)
        sheet.merge_cells(
            start_row=2,
            start_column=start_col,
            end_row=2,
            end_column=end_col
        )
        cell = sheet.cell(row=2, column=start_col, value=group_name)
        cell.font = group_font
        cell.fill = fill
        cell.alignment = center

        for col_index in range(start_col, end_col + 1):
            group_cell = sheet.cell(row=2, column=col_index)
            group_cell.fill = fill
            group_cell.border = border

    for col_index, column_name in enumerate(columns, start=1):
        cell = sheet.cell(row=3, column=col_index, value=column_name)
        cell.font = header_font
        cell.alignment = center
        cell.border = border

        if col_index <= 5:
            cell.fill = branch_fill
        elif col_index <= 8:
            cell.fill = total_fill
        elif col_index <= 15:
            cell.fill = alc_fill
        elif col_index <= 22:
            cell.fill = loan_fill
        elif col_index <= 29:
            cell.fill = arrear_fill
        elif col_index <= 36:
            cell.fill = outstanding_fill
        else:
            cell.fill = percent_fill

    for row_index, row_data in enumerate(rows, start=4):
        is_total = str(row_data[0]).strip().lower() == "grand total"
        for col_index, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.alignment = left if col_index in (2, 3, 4) else center

            if is_total:
                cell.font = header_font
                cell.fill = grand_total_fill

    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(max_col)}{len(rows) + 3}"
    sheet.row_dimensions[1].height = 22
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[3].height = 42

    for col_index, column_name in enumerate(columns, start=1):
        column_letter = get_column_letter(col_index)
        if col_index in (1, 5):
            width = 12
        elif col_index in (2, 3, 4):
            width = 24
        elif "%" in str(column_name):
            width = 18
        else:
            width = 15
        sheet.column_dimensions[column_letter].width = width


#------------------ LOGIN ------------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return render_template('Login.html')

    conn = None
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        username = request.form['username'].strip()
        password = request.form['password'].strip()
        remember_device = request.form.get('remember') == 'on'

        user = verify_user_password(cursor, username, password)
        conn.commit()
    except Exception as e:
        flash(f"Database connection failed: {e}", "error")
        return redirect('/login')
    finally:
        if conn:
            conn.close()

    if user == "inactive":
        log_activity("login_blocked", f"Inactive user attempted login: {username}")
        flash("Your account is waiting for admin approval or has been deactivated.", "error")
        return redirect('/login')

    if user == "locked":
        log_activity("login_blocked", f"Locked user attempted login: {username}")
        flash("Too many failed login attempts. Please reset password or contact admin.", "error")
        return redirect('/login')

    if user:
        registered_email = (user[1] or "").strip()

        if not registered_email:
            flash("Registered email address not found. Please contact admin.", "error")
            return redirect('/login')

        session['temp_user'] = user[0]
        session['temp_role'] = user[6] or "user"
        session['otp_email'] = registered_email
        session['remember_device'] = remember_device
        otp = str(random.randint(100000, 999999))
        session['otp'] = otp
        session['otp_expires_at'] = (datetime.now() + timedelta(minutes=5)).isoformat()
        sent, otp_message = send_otp_email(registered_email, otp)
        if not sent:
            print(f"OTP for {registered_email}: {otp}")
            sent = True
            otp_message = "Email OTP failed, so OTP printed in console"
        session['otp_message'] = otp_message
        session['otp_email_masked'] = mask_email(registered_email)
        log_activity("otp_sent", "Login OTP generated")
        return redirect('/otp')

    log_activity("login_failed", f"Invalid password for username: {username}")
    flash("Invalid username or password.", "error")
    return redirect('/login')


# ------------------ DB CONNECTION ------------------
def get_conn():
    db = session.get('db_config')
    if not db:
        return None

    try:
        return connect_sql_server(
            db['server'],
            db['database'],
            db['uid'],
            db['pwd']
        )
    except Exception as e:
        print("DB ERROR:", e)
        return None


def get_session_cache(cache_name):
    cache = session.get(cache_name)
    if not isinstance(cache, dict):
        cache = {}
    return cache


def cache_databases(server, databases):
    cache = get_session_cache('database_cache')
    cache[server] = databases
    session['database_cache'] = cache


def get_cached_databases(server):
    return get_session_cache('database_cache').get(server)


def table_cache_key(server, database):
    return f"{server}::{database}"


def cache_tables(server, database, tables):
    cache = get_session_cache('table_cache')
    cache[table_cache_key(server, database)] = tables
    session['table_cache'] = cache


def get_cached_tables(server, database):
    return get_session_cache('table_cache').get(table_cache_key(server, database))


def get_sql_credentials_from_session():
    credentials = session.get('sql_credentials')
    if not isinstance(credentials, dict):
        credentials = {}
    return {
        "uid": credentials.get("uid") or APP_DB_CONFIG["uid"],
        "pwd": credentials.get("pwd") or APP_DB_CONFIG["pwd"],
    }


def save_sql_credentials(uid, pwd):
    uid = (uid or "").strip()
    pwd = (pwd or "").strip()
    session['sql_credentials'] = {
        "uid": uid or APP_DB_CONFIG["uid"],
        "pwd": pwd or APP_DB_CONFIG["pwd"],
    }


def remember_sql_query(query_text):
    query_text = (query_text or "").strip()
    if not query_text:
        return

    history = session.get('sql_query_history', [])
    history = [q for q in history if q != query_text]
    history.insert(0, query_text)
    session['sql_query_history'] = history[:8]


def quote_sql_name(name):
    return "[" + str(name).replace("]", "]]") + "]"


def get_report_table_name(query_text):
    match = re.match(
        r"^\s*EXEC(?:UTE)?\s+(?:dbo\.)?"
        r"(usp_DPD_Report|usp_DPD_Report_1_to_30|usp_SIDBI_Prayaas_TopSheet)"
        r"\s+'((?:''|[^'])*)'\s*;?\s*$",
        query_text or "",
        re.IGNORECASE
    )
    if not match:
        return None, None

    proc_name = match.group(1).lower()
    table_name = match.group(2).replace("''", "'").strip()
    return proc_name, table_name


def get_safe_table_reference(table_name):
    parts = [part.strip() for part in (table_name or "").split(".") if part.strip()]
    if not parts or len(parts) > 2:
        raise ValueError("Only one-part or two-part table names are allowed.")
    if any(part in (".", "..") for part in parts):
        raise ValueError("Invalid table name.")

    if len(parts) == 1:
        parts.insert(0, "dbo")

    return ".".join(quote_sql_name(part) for part in parts)


def build_staff_wise_sidbi_query(table_name, one_to_thirty=False):
    table_ref = get_safe_table_reference(table_name)
    normal_last_bucket = "15-30 DPD" if one_to_thirty else "15-28 DPD"
    normal_last_limit = 30 if one_to_thirty else 28

    return f"""
WITH BaseData AS (
    SELECT
        LEFT(BranchID, 2) AS RegionID,
        DistrictName,
        BranchID,
        BranchName,
        CenterStaffID,
        StaffName,
        COUNT(DISTINCT LEFT(ClientLoanID, 14)) AS ALC,
        COUNT(ClientLoanID) AS Loan,
        SUM(CAST(CASE WHEN PrincipalInArrears > 0 THEN PrincipalInArrears ELSE 0 END AS NUMERIC)) AS ArrearAmt,
        SUM(CAST(PrincipalOutstanding AS FLOAT)) AS OutStanding,
        CASE
            WHEN DPDFlag = '0' AND max_dpd = 0 THEN '0 DPD'
            WHEN DPDFlag = '0' AND max_dpd BETWEEN 1 AND 14 THEN '1-14 DPD'
            WHEN DPDFlag = '0' AND max_dpd BETWEEN 15 AND {normal_last_limit} THEN '{normal_last_bucket}'
            WHEN DPDFlag = '1' AND max_dpd BETWEEN 1 AND 30 THEN '1-30 DPD S3'
            WHEN DPDFlag = '1' AND max_dpd BETWEEN 31 AND 60 THEN '31-60 DPD S3'
            WHEN DPDFlag = '1' AND max_dpd BETWEEN 61 AND 90 THEN '61-90 DPD S3'
            WHEN DPDFlag = '1' AND max_dpd > 90 THEN '>90 DPD S3'
        END AS Ageing
    FROM {table_ref}
    WHERE SUBSTRING(ClientLoanID, 16, 2) IN (
        'W6', '5X', 'H8', '6X', '6Y', 'Q1', 'Q2', 'Q3', 'Q4', 'Q5', 'Q6',
        '6D', '6E', '6F', '6G', '6H', '6I', '6J'
    )
    GROUP BY
        LEFT(BranchID, 2),
        DistrictName,
        BranchID,
        BranchName,
        CenterStaffID,
        StaffName,
        DPDFlag,
        max_dpd
)
SELECT
    CASE WHEN RegionID IS NULL THEN 'Grand Total' ELSE RegionID END AS RegionID,
    CASE WHEN RegionID IS NULL THEN NULL ELSE DistrictName END AS [Region Name],
    CASE WHEN RegionID IS NULL THEN NULL ELSE BranchName END AS [Branch Name],
    CASE WHEN RegionID IS NULL THEN NULL ELSE '[' + BranchID + ']-' + StaffName END AS [Branch / Staff Name],
    CASE WHEN RegionID IS NULL THEN NULL ELSE CenterStaffID END AS [Staff ID],
    SUM(ALC) AS ALC,
    SUM(Loan) AS Loan,
    SUM(OutStanding) AS OutStanding,
    SUM(CASE WHEN Ageing = '0 DPD' THEN ALC ELSE 0 END) AS [ALC 0 DPD],
    SUM(CASE WHEN Ageing = '1-14 DPD' THEN ALC ELSE 0 END) AS [ALC 1-14 DPD],
    SUM(CASE WHEN Ageing = '{normal_last_bucket}' THEN ALC ELSE 0 END) AS [ALC {normal_last_bucket}],
    SUM(CASE WHEN Ageing = '1-30 DPD S3' THEN ALC ELSE 0 END) AS [ALC 1-30 DPD S3],
    SUM(CASE WHEN Ageing = '31-60 DPD S3' THEN ALC ELSE 0 END) AS [ALC 31-60 DPD S3],
    SUM(CASE WHEN Ageing = '61-90 DPD S3' THEN ALC ELSE 0 END) AS [ALC 61-90 DPD S3],
    SUM(CASE WHEN Ageing = '>90 DPD S3' THEN ALC ELSE 0 END) AS [ALC >90 DPD S3],
    SUM(CASE WHEN Ageing = '0 DPD' THEN Loan ELSE 0 END) AS [Loan 0 DPD],
    SUM(CASE WHEN Ageing = '1-14 DPD' THEN Loan ELSE 0 END) AS [Loan 1-14 DPD],
    SUM(CASE WHEN Ageing = '{normal_last_bucket}' THEN Loan ELSE 0 END) AS [Loan {normal_last_bucket}],
    SUM(CASE WHEN Ageing = '1-30 DPD S3' THEN Loan ELSE 0 END) AS [Loan 1-30 DPD S3],
    SUM(CASE WHEN Ageing = '31-60 DPD S3' THEN Loan ELSE 0 END) AS [Loan 31-60 DPD S3],
    SUM(CASE WHEN Ageing = '61-90 DPD S3' THEN Loan ELSE 0 END) AS [Loan 61-90 DPD S3],
    SUM(CASE WHEN Ageing = '>90 DPD S3' THEN Loan ELSE 0 END) AS [Loan >90 DPD S3],
    SUM(CASE WHEN Ageing = '0 DPD' THEN ArrearAmt ELSE 0 END) AS [ArrearAmt 0 DPD],
    SUM(CASE WHEN Ageing = '1-14 DPD' THEN ArrearAmt ELSE 0 END) AS [ArrearAmt 1-14 DPD],
    SUM(CASE WHEN Ageing = '{normal_last_bucket}' THEN ArrearAmt ELSE 0 END) AS [ArrearAmt {normal_last_bucket}],
    SUM(CASE WHEN Ageing = '1-30 DPD S3' THEN ArrearAmt ELSE 0 END) AS [ArrearAmt 1-30 DPD S3],
    SUM(CASE WHEN Ageing = '31-60 DPD S3' THEN ArrearAmt ELSE 0 END) AS [ArrearAmt 31-60 DPD S3],
    SUM(CASE WHEN Ageing = '61-90 DPD S3' THEN ArrearAmt ELSE 0 END) AS [ArrearAmt 61-90 DPD S3],
    SUM(CASE WHEN Ageing = '>90 DPD S3' THEN ArrearAmt ELSE 0 END) AS [ArrearAmt >90 DPD S3],
    SUM(CASE WHEN Ageing = '0 DPD' THEN OutStanding ELSE 0 END) AS [Outstanding 0 DPD],
    SUM(CASE WHEN Ageing = '1-14 DPD' THEN OutStanding ELSE 0 END) AS [Outstanding 1-14 DPD],
    SUM(CASE WHEN Ageing = '{normal_last_bucket}' THEN OutStanding ELSE 0 END) AS [Outstanding {normal_last_bucket}],
    SUM(CASE WHEN Ageing = '1-30 DPD S3' THEN OutStanding ELSE 0 END) AS [Outstanding 1-30 DPD S3],
    SUM(CASE WHEN Ageing = '31-60 DPD S3' THEN OutStanding ELSE 0 END) AS [Outstanding 31-60 DPD S3],
    SUM(CASE WHEN Ageing = '61-90 DPD S3' THEN OutStanding ELSE 0 END) AS [Outstanding 61-90 DPD S3],
    SUM(CASE WHEN Ageing = '>90 DPD S3' THEN OutStanding ELSE 0 END) AS [Outstanding >90 DPD S3],
    CAST(CASE WHEN SUM(OutStanding) = 0 THEN 0 ELSE ROUND((SUM(CASE WHEN Ageing IN ('1-30 DPD S3', '31-60 DPD S3', '61-90 DPD S3', '>90 DPD S3') THEN OutStanding ELSE 0 END) * 100.0) / SUM(OutStanding), 3) END AS VARCHAR) + '%' AS [% Outstanding >30 DPD],
    CAST(CASE WHEN SUM(OutStanding) = 0 THEN 0 ELSE ROUND((SUM(CASE WHEN Ageing IN ('61-90 DPD S3', '>90 DPD S3') THEN OutStanding ELSE 0 END) * 100.0) / SUM(OutStanding), 3) END AS VARCHAR) + '%' AS [% Outstanding >60 DPD],
    CAST(CASE WHEN SUM(OutStanding) = 0 THEN 0 ELSE ROUND((SUM(CASE WHEN Ageing = '>90 DPD S3' THEN OutStanding ELSE 0 END) * 100.0) / SUM(OutStanding), 3) END AS VARCHAR) + '%' AS [% Outstanding >90 DPD]
FROM BaseData
GROUP BY GROUPING SETS (
    (RegionID, DistrictName, BranchName, BranchID, CenterStaffID, StaffName),
    ()
)
ORDER BY
    CASE WHEN RegionID IS NULL THEN 1 ELSE 0 END,
    RegionID,
    BranchName,
    CenterStaffID
"""


def build_sidbi_top_sheet_query(table_name):
    table_ref = get_safe_table_reference(table_name)
    return f"""
WITH BaseData AS (
    SELECT
        COUNT(DISTINCT LEFT(ClientLoanID, 14)) AS ALC,
        COUNT(ClientLoanID) AS Loan,
        SUM(CAST(CASE WHEN PrincipalInArrears > 0 THEN PrincipalInArrears ELSE 0 END AS NUMERIC)) AS ArrearAmt,
        SUM(CAST(PrincipalOutstanding AS FLOAT)) AS OutStanding,
        DPDFlag,
        max_dpd
    FROM {table_ref}
    WHERE SUBSTRING(ClientLoanID, 16, 2) IN (
        'W6', '5X', 'H8', '6X', '6Y', 'Q1', 'Q2', 'Q3', 'Q4', 'Q5', 'Q6',
        '6D', '6E', '6F', '6G', '6H', '6I', '6J'
    )
    GROUP BY DPDFlag, max_dpd
),
BucketData AS (
    SELECT
        CASE
            WHEN DPDFlag = '0' AND max_dpd = 0 THEN '0 DPD'
            WHEN DPDFlag = '0' AND max_dpd BETWEEN 1 AND 14 THEN '1-14 DPD'
            WHEN DPDFlag = '0' AND max_dpd BETWEEN 15 AND 28 THEN '15-28 DPD'
            WHEN DPDFlag = '1' AND max_dpd BETWEEN 1 AND 30 THEN '1-30 DPD S3'
            WHEN DPDFlag = '1' AND max_dpd BETWEEN 31 AND 60 THEN '31-60 DPD S3'
            WHEN DPDFlag = '1' AND max_dpd BETWEEN 61 AND 90 THEN '61-90 DPD S3'
            WHEN DPDFlag = '1' AND max_dpd > 90 THEN '>90 DPD S3'
        END AS LeftBucket,
        CASE
            WHEN DPDFlag = '0' AND max_dpd = 0 THEN '0 DPD'
            WHEN DPDFlag = '0' AND max_dpd BETWEEN 1 AND 14 THEN '1-14 DPD'
            WHEN DPDFlag = '0' AND max_dpd BETWEEN 15 AND 30 THEN '15-30 DPD'
            WHEN DPDFlag = '1' AND max_dpd BETWEEN 1 AND 30 THEN '1-30 DPD S3'
            WHEN DPDFlag = '1' AND max_dpd BETWEEN 31 AND 60 THEN '31-60 DPD S3'
            WHEN DPDFlag = '1' AND max_dpd BETWEEN 61 AND 90 THEN '61-90 DPD S3'
            WHEN DPDFlag = '1' AND max_dpd > 90 THEN '>90 DPD S3'
        END AS RightBucket,
        ALC,
        Loan,
        ArrearAmt,
        OutStanding
    FROM BaseData
),
LeftSide AS (
    SELECT 1 AS SortOrder, '0 DPD' AS Bucket
    UNION ALL SELECT 2, '1-14 DPD'
    UNION ALL SELECT 3, '15-28 DPD'
    UNION ALL SELECT 4, '1-30 DPD S3'
    UNION ALL SELECT 5, '31-60 DPD S3'
    UNION ALL SELECT 6, '61-90 DPD S3'
    UNION ALL SELECT 7, '>90 DPD S3'
    UNION ALL SELECT 8, 'Grand Total'
),
RightSide AS (
    SELECT 1 AS SortOrder, '0 DPD' AS Bucket
    UNION ALL SELECT 2, '1-14 DPD'
    UNION ALL SELECT 3, '15-30 DPD'
    UNION ALL SELECT 4, '1-30 DPD S3'
    UNION ALL SELECT 5, '31-60 DPD S3'
    UNION ALL SELECT 6, '61-90 DPD S3'
    UNION ALL SELECT 7, '>90 DPD S3'
    UNION ALL SELECT 8, 'Grand Total'
),
LeftAgg AS (
    SELECT
        L.SortOrder,
        L.Bucket,
        SUM(CASE WHEN L.Bucket = 'Grand Total' OR B.LeftBucket = L.Bucket THEN B.ALC ELSE 0 END) AS ALC,
        SUM(CASE WHEN L.Bucket = 'Grand Total' OR B.LeftBucket = L.Bucket THEN B.Loan ELSE 0 END) AS Loan,
        SUM(CASE WHEN L.Bucket = 'Grand Total' OR B.LeftBucket = L.Bucket THEN B.ArrearAmt ELSE 0 END) AS ArrearAmt,
        SUM(CASE WHEN L.Bucket = 'Grand Total' OR B.LeftBucket = L.Bucket THEN B.OutStanding ELSE 0 END) AS OutStanding
    FROM LeftSide L
    LEFT JOIN BucketData B ON L.Bucket = 'Grand Total' OR B.LeftBucket = L.Bucket
    GROUP BY L.SortOrder, L.Bucket
),
RightAgg AS (
    SELECT
        R.SortOrder,
        R.Bucket,
        SUM(CASE WHEN R.Bucket = 'Grand Total' OR B.RightBucket = R.Bucket THEN B.ALC ELSE 0 END) AS ALC,
        SUM(CASE WHEN R.Bucket = 'Grand Total' OR B.RightBucket = R.Bucket THEN B.Loan ELSE 0 END) AS Loan,
        SUM(CASE WHEN R.Bucket = 'Grand Total' OR B.RightBucket = R.Bucket THEN B.ArrearAmt ELSE 0 END) AS ArrearAmt,
        SUM(CASE WHEN R.Bucket = 'Grand Total' OR B.RightBucket = R.Bucket THEN B.OutStanding ELSE 0 END) AS OutStanding
    FROM RightSide R
    LEFT JOIN BucketData B ON R.Bucket = 'Grand Total' OR B.RightBucket = R.Bucket
    GROUP BY R.SortOrder, R.Bucket
)
SELECT
    L.Bucket AS Bucket,
    L.ALC AS ALC,
    L.Loan AS Loan,
    L.ArrearAmt AS ArrearAmt,
    L.OutStanding AS OutStanding,
    '' AS [ ],
    R.Bucket AS Bucket,
    R.ALC AS ALC,
    R.Loan AS Loan,
    R.ArrearAmt AS ArrearAmt,
    R.OutStanding AS OutStanding
FROM LeftAgg L
JOIN RightAgg R ON R.SortOrder = L.SortOrder
ORDER BY L.SortOrder
"""


def get_backend_report_query(query_text):
    proc_name, table_name = get_report_table_name(query_text)
    if not proc_name:
        return query_text
    if proc_name == "usp_sidbi_prayaas_topsheet":
        return build_sidbi_top_sheet_query(table_name)
    return build_staff_wise_sidbi_query(
        table_name,
        one_to_thirty=proc_name == "usp_dpd_report_1_to_30"
    )


def fetch_cursor_result(cursor):
    while cursor.description is None:
        if not cursor.nextset():
            return None

    columns = [col[0] for col in cursor.description]
    rows = [list(r) for r in cursor.fetchall()]
    return {
        "columns": columns,
        "rows": rows
    }


def execute_sql_text(cursor, query_text):
    require_safe_sql(query_text)
    backend_query = get_backend_report_query(query_text)
    cursor.execute(backend_query)
    return fetch_cursor_result(cursor)


def run_sql_with_metrics(cursor, query_text):
    started = time.perf_counter()
    query_result = execute_sql_text(cursor, query_text)
    elapsed_ms = int((time.perf_counter() - started) * 1000)
    rows_count = len(query_result["rows"]) if query_result else 0
    return query_result, {
        "elapsed_ms": elapsed_ms,
        "rows_count": rows_count,
        "status": "Success"
    }


def get_table_preview_query(table_name):
    return f"SELECT TOP 100 * FROM {get_safe_table_reference(table_name)}"


def get_table_count_query(table_name):
    return f"SELECT COUNT(*) AS RowCount FROM {get_safe_table_reference(table_name)}"


def get_table_columns_result(cursor, table_name):
    parts = [part.strip() for part in (table_name or "").split(".") if part.strip()]
    schema_name = parts[0] if len(parts) == 2 else "dbo"
    object_name = parts[-1] if parts else ""
    cursor.execute(
        """
        SELECT COLUMN_NAME, DATA_TYPE, IS_NULLABLE
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA=? AND TABLE_NAME=?
        ORDER BY ORDINAL_POSITION
        """,
        (schema_name, object_name)
    )
    return {
        "columns": ["Column", "Type", "Nullable"],
        "rows": [list(row) for row in cursor.fetchall()]
    }


def validate_sidbi_table(cursor, table_name):
    required_columns = [
        "BranchID", "DistrictName", "BranchName", "CenterStaffID",
        "StaffName", "ClientLoanID", "PrincipalInArrears",
        "PrincipalOutstanding", "DPDFlag", "max_dpd"
    ]
    existing_result = get_table_columns_result(cursor, table_name)
    existing_columns = {row[0] for row in existing_result["rows"]}
    missing_columns = [col for col in required_columns if col not in existing_columns]
    status = "Ready" if not missing_columns else "Missing columns"
    return {
        "columns": ["Status", "Missing Columns", "Required Columns"],
        "rows": [[status, ", ".join(missing_columns) or "None", ", ".join(required_columns)]]
    }


def export_rows_as_csv(columns, rows):
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    writer.writerows(rows)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=query_result.csv"}
    )


# ------------------ HOME ------------------
@app.route('/')
def home():
    return redirect('/login')


# ------------------ DB CONFIG ------------------
@app.route('/db_config')
@login_required
def db_config():
    return render_template('db_config.html')


@app.route('/get_databases', methods=['POST'])
@login_required
def get_databases():
    server = request.form.get('server')
    uid = request.form.get('uid')
    pwd = request.form.get('pwd')

    conn = None
    try:
        conn = connect_sql_server(server, uid=uid, pwd=pwd)

        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sys.databases")

        databases = [row[0] for row in cursor.fetchall()]
        return {"databases": databases}

    except Exception as e:
        return {"error": str(e)}
    finally:
        if conn:
            conn.close()


# ------------------ TABLE LIST ------------------
@app.route('/get_tables')
@login_required
def get_tables():
    conn = get_conn()
    if not conn:
        return {"tables": []}

    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT TABLE_NAME 
            FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_TYPE='BASE TABLE'
            ORDER BY TABLE_NAME
        """)

        tables = [row[0] for row in cursor.fetchall()]
        return {"tables": tables}
    finally:
        conn.close()


# ------------------ CONNECT DB ------------------
@app.route('/connect_db', methods=['POST'])
@login_required
def connect_db():
    session['db_config'] = {
        "server": request.form['server'],
        "database": request.form['database'],
        "uid": request.form['uid'],
        "pwd": request.form['pwd']
    }
    flash("Database connected successfully.", "success")
    return redirect('/dashboard')


# ------------------ REGISTER Page------------------
@app.route('/register')
def register():
    return render_template('register.html')

# ------------------ REGISTER USER------------------
@app.route('/register_user', methods=['POST'])
def register_user():

    conn = None
    try:
        conn = get_app_conn()
        cursor = conn.cursor()

        username = request.form['username'].strip()
        password = request.form['password'].strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        email = request.form['email'].strip()
        mobile = normalize_mobile(request.form.get('mobile', '').strip())

        if not username or not password:
            flash("Username and password are required.", "error")
            return redirect('/register')

        if password != confirm_password:
            flash("Password and confirm password must match.", "error")
            return redirect('/register')

        if len(password) < 6:
            flash("Password must be at least 6 characters.", "error")
            return redirect('/register')

        if not mobile:
            flash("Enter a valid 10 digit mobile number.", "error")
            return redirect('/register')

        cursor.execute("SELECT 1 FROM users WHERE username=?", (username,))
        existing = cursor.fetchone()

        if existing:
            flash("User already exists.", "error")
            return redirect('/register')

        cursor.execute("SELECT COUNT(*) FROM dbo.users")
        role = "admin" if cursor.fetchone()[0] == 0 else "user"
        is_active = 1 if role == "admin" else 0

        cursor.execute(
            """
            INSERT INTO users (username, password, password_hash, email, mobile, profile_pic, role, is_active)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (username, "", generate_password_hash(password), email, mobile, 'Default.png', role, is_active)
        )

        conn.commit()
        log_activity("user_registered", f"Registered username: {username}; active={is_active}")
        if role == "admin":
            flash("Admin account created successfully. Please login.", "success")
        else:
            flash("Account created successfully. Please wait for admin approval before login.", "success")
        return redirect('/login')
    except Exception as e:
        flash(f"Registration failed: {e}", "error")
        return redirect('/register')
    finally:
        if conn:
            conn.close()


# ------------------ FORGOT PASSWORD ------------------
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'GET':
        return render_template('forgot_password.html')

    username = request.form.get('username', '').strip()
    identity = request.form.get('identity', '').strip()
    new_password = request.form.get('new_password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()

    if not username or not identity or not new_password or not confirm_password:
        return render_template(
            'forgot_password.html',
            error='All fields are required'
        )

    if new_password != confirm_password:
        return render_template(
            'forgot_password.html',
            error='New password and confirm password do not match'
        )

    conn = None
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT 1
            FROM users
            WHERE username=? AND (email=? OR mobile=?)
            """,
            (username, identity, identity)
        )
        user = cursor.fetchone()

        if not user:
            return render_template(
                'forgot_password.html',
                error='Username with this email/mobile was not found'
            )

        cursor.execute(
            "UPDATE users SET password='', password_hash=?, failed_attempts=0 WHERE username=?",
            (generate_password_hash(new_password), username)
        )
        conn.commit()
        log_activity("password_reset", f"Password reset for username: {username}")

        return render_template(
            'forgot_password.html',
            success='Password reset successfully. You can login now.'
        )

    except Exception as e:
        return render_template(
            'forgot_password.html',
            error=f'Password reset failed: {e}'
        )
    finally:
        if conn:
            conn.close()


# ------------------ OTP ------------------
@app.route('/otp')
def otp_page():
    if not session.get('otp'):
        return redirect('/login')
    return render_template(
        'otp.html',
        otp_email=session.get('otp_email_masked', ''),
        otp_message=session.get('otp_message', '')
    )


@app.route('/resend_otp', methods=['POST'])
def resend_otp():
    if not session.get('temp_user') or not session.get('otp_email'):
        return redirect('/login')

    otp = str(random.randint(100000, 999999))
    session['otp'] = otp
    session['otp_expires_at'] = (datetime.now() + timedelta(minutes=5)).isoformat()
    sent, otp_message = send_otp_email(session.get('otp_email'), otp)
    if not sent:
        print(f"OTP for {session.get('otp_email')}: {otp}")
        sent = True
        otp_message = "Email OTP failed, so OTP printed in console"
    session['otp_message'] = otp_message
    session['otp_email_masked'] = mask_email(session.get('otp_email'))

    return redirect('/otp')


@app.route('/verify_otp', methods=['POST'])
def verify_otp():
    expires_at = session.get("otp_expires_at")
    if expires_at:
        try:
            if datetime.fromisoformat(expires_at) < datetime.now():
                session.pop('otp', None)
                session.pop('otp_expires_at', None)
                log_activity("otp_expired", "OTP verification attempted after expiry")
                flash("OTP expired. Please login again or request a new OTP.", "error")
                return redirect('/login')
        except ValueError:
            pass

    if request.form.get('otp') == session.get('otp'):
        session.permanent = bool(session.get('remember_device'))
        session['logged_in'] = True
        session['username'] = session.get('temp_user')
        session['role'] = session.get('temp_role') or "user"
        session.pop('otp', None)
        session.pop('otp_expires_at', None)
        session.pop('otp_mobile', None)
        session.pop('otp_mobile_masked', None)
        session.pop('otp_email', None)
        session.pop('otp_email_masked', None)
        session.pop('otp_message', None)
        session.pop('remember_device', None)
        session.pop('temp_role', None)
        session.pop('temp_user', None)
        conn = None
        try:
            conn = get_app_conn()
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE dbo.users SET last_login=GETDATE() WHERE username=?",
                (session.get('username'),)
            )
            conn.commit()
        except Exception as e:
            print(f"Last login update failed: {e}")
        finally:
            if conn:
                conn.close()
        log_activity("login_success", "OTP verified")
        return redirect('/dashboard')

    log_activity("otp_failed", "Wrong OTP entered")
    flash("Wrong OTP. Please enter the latest 6 digit code.", "error")
    return redirect('/otp')


# ------------------ DASHBOARD ------------------
@app.route('/dashboard')
@login_required
def dashboard():
    now = datetime.now()

    wish = (
        "Good Morning" if now.hour < 12 else
        "Good Afternoon" if now.hour < 17 else
        "Good Evening"
    )

    return render_template(
        "dashboard.html",
        username=session.get('username'),
        role=session.get('role', 'user'),
        wish=wish,
        current_date=now.strftime("%d %B %Y"),
        stats=fetch_dashboard_stats()
    )
# ------------------ PROFILE ------------------
@app.route('/profile')
@login_required
def profile():
    conn = None
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT profile_pic FROM users WHERE username=?",
            (session.get('username'),)
        )
        row = cursor.fetchone()
    except Exception as e:
        flash(f"Database connection failed: {e}", "error")
        row = None
    finally:
        if conn:
            conn.close()

    profile_pic = row[0] if row and row[0] else "Default.png"

    return render_template(
        'profile.html',
        username=session.get('username'),
        profile_pic=profile_pic
    )
@app.route('/add_student')
@login_required
def add_student():
    return render_template('Add_student.html')



# ------------------ ADD STUDENT ------------------
@app.route('/save_student', methods=['POST'])
@login_required
def save_student():
    conn = None
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO dbo.students (name, age, course) VALUES (?, ?, ?)",
            (request.form['name'], request.form['age'], request.form['course'])
        )
        conn.commit()
        log_activity("student_added", f"Added student: {request.form['name']}")
        flash("Student added successfully.", "success")

    except Exception as e:
        flash(f"Database error: {e}", "error")
        return redirect('/add_student')

    finally:
        if conn:
            conn.close()

    return redirect('/view_students')


# ------------------ VIEW STUDENTS ------------------
@app.route('/view_students')
@login_required
def view_students():
    search = request.args.get("q", "").strip()
    course = request.args.get("course", "").strip()
    conn = None
    students = []
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        sql = "SELECT * FROM students WHERE 1=1"
        params = []
        if search:
            sql += " AND (name LIKE ? OR course LIKE ?)"
            params.extend([f"%{search}%", f"%{search}%"])
        if course:
            sql += " AND course LIKE ?"
            params.append(f"%{course}%")
        sql += " ORDER BY id DESC"
        cursor.execute(sql, tuple(params))
        students = cursor.fetchall()

    except Exception as e:
        flash(f"Could not load students: {e}", "error")

    finally:
        if conn:
            conn.close()

        return render_template(
            'Students.html',
            students=students,
            search=search,
            course=course
        )
# ------------------ DELETE STUDENT ------------------
@app.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete_student(id):

    conn = None
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM students WHERE id=?", (id,))
        conn.commit()
        log_activity("student_deleted", f"Deleted student id: {id}")
        flash("Student deleted successfully.", "success")

    except Exception as e:
        flash(f"Delete failed: {e}", "error")

    finally:
        if conn:
            conn.close()

    return redirect('/view_students')


@app.route('/import_students', methods=['POST'])
@login_required
def import_students():
    upload = request.files.get("student_file")
    if not upload or not upload.filename:
        flash("Please choose an Excel file to import.", "warning")
        return redirect('/view_students')

    conn = None
    imported = 0
    try:
        wb = load_workbook(upload, read_only=True, data_only=True)
        sheet = wb.active
        conn = get_app_conn()
        cursor = conn.cursor()

        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not row or index == 1 and str(row[0]).strip().lower() in ("name", "student name"):
                continue
            name = str(row[0] or "").strip()
            age = row[1] if len(row) > 1 else ""
            course = str(row[2] or "").strip() if len(row) > 2 else ""
            if not name or not age or not course:
                continue
            cursor.execute(
                "INSERT INTO dbo.students (name, age, course) VALUES (?, ?, ?)",
                (name, age, course)
            )
            imported += 1

        conn.commit()
        log_activity("students_imported", f"Imported {imported} students from Excel")
        flash(f"Imported {imported} students from Excel.", "success")
    except Exception as e:
        flash(f"Import failed: {e}", "error")
    finally:
        if conn:
            conn.close()

    return redirect('/view_students')


@app.route('/download_student_template')
@login_required
def download_student_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["Name", "Age", "Course"])
    ws.append(["Sample Student", 21, "Computer"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="student_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ------------------ DELETE ACCOUNT ------------------
@app.route('/delete_account', methods=['POST'])
@login_required
def delete_account():
    conn = None
    try:
        conn = get_app_conn()
        cursor = conn.cursor()

        cursor.execute(
            "DELETE FROM users WHERE username=?",
            (session.get('username'),)
        )
        conn.commit()
        log_activity("account_deleted", "User deleted own account")
    except Exception as e:
        flash(f"Account delete failed: {e}", "error")
        return redirect('/profile')
    finally:
        if conn:
            conn.close()

    session.clear()
    flash("Account deleted successfully.", "success")
    return redirect('/login')


# ------------------ SETTINGS ------------------
@app.route('/settings')
@login_required
def settings():
    profile_pic = "Default.png"
    email = ""
    image_dir = os.path.join(app.root_path, "static", "Images")
    profile_images = []

    if os.path.isdir(image_dir):
        profile_images = sorted(
            name for name in os.listdir(image_dir)
            if name.lower().endswith((".png", ".jpg", ".jpeg", ".gif", ".webp"))
        )

    conn = None
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT profile_pic, email FROM dbo.users WHERE username=?",
            (session.get("username"),)
        )
        row = cursor.fetchone()
        if row and row[0]:
            profile_pic = row[0]
        if row and row[1]:
            email = row[1]
    except Exception as e:
        flash(f"Could not load profile settings: {e}", "error")
    finally:
        if conn:
            conn.close()

    if profile_pic not in profile_images and profile_images:
        profile_images.insert(0, profile_pic)

    return render_template(
        'Settings.html',
        username=session.get("username"),
        email=email,
        profile_pic=profile_pic,
        profile_images=profile_images or ["Default.png"],
    )


@app.route('/activity_logs')
@login_required
def activity_logs():
    conn = None
    logs = []
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT TOP 100 username, action, details, ip_address, created_at
            FROM dbo.activity_logs
            ORDER BY id DESC
        """)
        logs = cursor.fetchall()
    except Exception as e:
        flash(f"Activity log error: {e}", "error")
    finally:
        if conn:
            conn.close()

    return render_template("activity_logs.html", logs=logs)


@app.route('/download_activity_logs_csv')
@login_required
def download_activity_logs_csv():
    conn = None
    rows = []
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT TOP 100 username, action, details, ip_address, created_at
            FROM dbo.activity_logs
            ORDER BY id DESC
        """)
        rows = cursor.fetchall()
    except Exception as e:
        flash(f"Activity log export failed: {e}", "error")
        return redirect('/activity_logs')
    finally:
        if conn:
            conn.close()

    log_activity("activity_logs_exported", "Downloaded activity logs CSV")
    return export_rows_as_csv(["User", "Action", "Details", "IP", "Time"], rows)


@app.route('/report_history')
@login_required
def report_history():
    conn = None
    reports = []
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT TOP 100 report_name, rows_count, sheets_count, created_at, username
            FROM dbo.report_history
            ORDER BY id DESC
        """)
        reports = cursor.fetchall()
    except Exception as e:
        flash(f"Report history error: {e}", "error")
    finally:
        if conn:
            conn.close()

    return render_template("report_history.html", reports=reports)


@app.route('/download_report_history_csv')
@login_required
def download_report_history_csv():
    conn = None
    rows = []
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT TOP 100 report_name, rows_count, sheets_count, created_at, username
            FROM dbo.report_history
            ORDER BY id DESC
        """)
        rows = cursor.fetchall()
    except Exception as e:
        flash(f"Report history export failed: {e}", "error")
        return redirect('/report_history')
    finally:
        if conn:
            conn.close()

    log_activity("report_history_exported", "Downloaded report history CSV")
    return export_rows_as_csv(["Report", "Rows", "Sheets", "Time", "User"], rows)


@app.route('/admin/users')
@admin_required
def admin_users():
    conn = None
    users = []
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT username, email, mobile, role, is_active, last_login, failed_attempts
            FROM dbo.users
            ORDER BY is_active ASC, username
        """)
        users = cursor.fetchall()
    except Exception as e:
        flash(f"User management error: {e}", "error")
    finally:
        if conn:
            conn.close()

    return render_template("users.html", users=users)


@app.route('/admin/users/export_csv')
@admin_required
def admin_users_export_csv():
    conn = None
    rows = []
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT username, email, mobile, role, is_active, last_login, failed_attempts
            FROM dbo.users
            ORDER BY is_active ASC, username
        """)
        rows = cursor.fetchall()
    except Exception as e:
        flash(f"User export failed: {e}", "error")
        return redirect('/admin/users')
    finally:
        if conn:
            conn.close()

    log_activity("users_exported", "Downloaded users CSV")
    return export_rows_as_csv(["Username", "Email", "Mobile", "Role", "Active", "Last Login", "Fails"], rows)


@app.route('/admin/users/update', methods=['POST'])
@admin_required
def admin_update_user():
    username = request.form.get("username", "").strip()
    role = request.form.get("role", "user").strip()
    is_active = 1 if request.form.get("is_active") == "1" else 0
    if role not in ("admin", "user", "sql_user"):
        role = "user"
    if username == session.get("username") and (role != "admin" or is_active == 0):
        flash("You cannot remove your own admin access or deactivate your own account.", "error")
        return redirect('/admin/users')

    conn = None
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE dbo.users SET role=?, is_active=?, failed_attempts=0 WHERE username=?",
            (role, is_active, username)
        )
        conn.commit()
        log_activity("user_updated", f"Updated user {username}: role={role}, active={is_active}")
        flash(f"User {username} updated successfully.", "success")
    except Exception as e:
        flash(f"User update failed: {e}", "error")
    finally:
        if conn:
            conn.close()

    return redirect('/admin/users')


# ------------------ UPDATE PROFILE ------------------
@app.route('/update_profile', methods=['POST'])
@login_required
def update_profile():
    conn = None
    try:
        current_username = session.get('username')
        conn = get_app_conn()
        new_username = request.form['username'].strip()
        new_email = request.form.get('email', '').strip()
        new_pic = request.form['profile_pic'].strip()

        if not new_username:
            flash("Username is required.", "error")
            return redirect('/settings')

        if "@" not in new_email:
            flash("Enter a valid email address.", "error")
            return redirect('/settings')

        cursor = conn.cursor()
        cursor.execute(
            "SELECT 1 FROM dbo.users WHERE username=? AND username<>?",
            (new_username, current_username)
        )
        if cursor.fetchone():
            flash("This username is already used by another account.", "error")
            return redirect('/settings')

        cursor.execute(
            "SELECT 1 FROM dbo.users WHERE email=? AND username<>?",
            (new_email, current_username)
        )
        if cursor.fetchone():
            flash("This email is already used by another account.", "error")
            return redirect('/settings')

        cursor.execute(
            "UPDATE users SET username=?, email=?, profile_pic=? WHERE username=?",
            (new_username, new_email, new_pic, current_username)
        )
        conn.commit()
        log_activity("profile_updated", f"Profile updated for {new_username}")
    except Exception as e:
        flash(f"Database connection failed: {e}", "error")
        return redirect('/settings')
    finally:
        if conn:
            conn.close()

    session['username'] = new_username
    flash("Profile updated successfully.", "success")

    return redirect('/profile')


# ------------------ CHANGE PASSWORD ------------------
@app.route('/change_password', methods=['POST'])
@login_required
def change_password():
    conn = None
    try:
        conn = get_app_conn()
        old_pass = request.form['old_password']
        new_pass = request.form['new_password']
        re_pass = request.form.get('re_password', '')

        if not re_pass or re_pass != new_pass:
            flash("New Password and Re-Enter Password must match.", "error")
            return redirect('/settings')

        if len(new_pass) < 6:
            flash("New password must be at least 6 characters.", "error")
            return redirect('/settings')

        cursor = conn.cursor()

        user = verify_user_password(cursor, session.get('username'), old_pass)

        if user and user not in ("inactive", "locked"):
            cursor.execute(
                "UPDATE users SET password='', password_hash=?, failed_attempts=0 WHERE username=?",
                (generate_password_hash(new_pass), session.get('username'))
            )
            conn.commit()
            log_activity("password_changed", "User changed password")
            flash("Password changed successfully.", "success")
            return redirect('/settings')
        flash("Old password is incorrect.", "error")
        return redirect('/settings')
    except Exception as e:
        flash(f"DB error: {e}", "error")
        return redirect('/settings')
    finally:
        if conn:
            conn.close()

@app.route('/sql', methods=['GET', 'POST'])
@login_required
def sql_runner():
    result = None
    error = None
    result_meta = None
    connection_status = None

    databases = []
    tables = []

    selected_server = ""
    selected_database = ""
    selected_table = ""
    sql_credentials = get_sql_credentials_from_session()
    selected_uid = sql_credentials["uid"]
    selected_pwd = sql_credentials["pwd"]

    saved_db_config = session.get("db_config")
    if saved_db_config:
        selected_server = saved_db_config.get("server", "")
        selected_database = saved_db_config.get("database", "")
        selected_uid = saved_db_config.get("uid", selected_uid).strip()
        selected_pwd = saved_db_config.get("pwd", selected_pwd).strip()

    query_text = ""
    dpd_table_name = ""

    servers = DEFAULT_SERVERS

    if request.method == 'POST':

        action = request.form.get('action')

        selected_server = request.form.get('server', '')
        selected_database = request.form.get('database', '')
        selected_table = request.form.get('table', '')
        selected_uid = request.form.get('uid', selected_uid).strip()
        selected_pwd = request.form.get('pwd', selected_pwd).strip()
        save_sql_credentials(selected_uid, selected_pwd)

        query_text = request.form.get('query', '').strip()
        compare_query = request.form.get('compare_query', '').strip()
        multi_query = request.form.get('multi_query', '').strip()
        dpd_table_name = request.form.get('dpd_table_name', '').strip()
        report_month = request.form.get('report_month', '').strip()

        # ================= SERVER CONNECT =================

        if selected_server:

            cached_databases = get_cached_databases(selected_server)

            if cached_databases and action != "connect_server":
                databases = cached_databases
            else:
                server_conn = None
                try:

                    server_conn = connect_sql_server(
                        selected_server,
                        uid=selected_uid,
                        pwd=selected_pwd
                    )

                    server_cursor = server_conn.cursor()

                    server_cursor.execute("""
                        SELECT name
                        FROM sys.databases
                        ORDER BY name
                    """)

                    databases = [row[0] for row in server_cursor.fetchall()]
                    cache_databases(selected_server, databases)

                except Exception as e:

                    error = f"Server connection failed: {e}"

                finally:

                    if server_conn:

                        server_conn.close()

        # ================= DATABASE CONNECT =================

        if selected_server and selected_database:

            conn = None
            try:

                conn = connect_sql_server(
                    selected_server,
                    selected_database,
                    selected_uid,
                    selected_pwd
                )

                cursor = conn.cursor()
                cursor.arraysize = 1000

                # ================= TABLE LIST =================

                cached_tables = get_cached_tables(
                    selected_server,
                    selected_database
                )

                if cached_tables:
                    tables = cached_tables
                else:
                    cursor.execute("""
                        SELECT TABLE_NAME
                        FROM INFORMATION_SCHEMA.TABLES
                        WHERE TABLE_TYPE='BASE TABLE'
                        ORDER BY TABLE_NAME
                    """)

                    tables = [row[0] for row in cursor.fetchall()]
                    cache_tables(selected_server, selected_database, tables)

                connection_status = {
                    "server": selected_server,
                    "database": selected_database,
                    "tables_count": len(tables),
                    "user": selected_uid
                }

                # ================= RUN QUERY =================

                if action == "dpd_report":

                    dpd_table_name = dpd_table_name or "BookDebt05032026"
                    query_text = f"EXEC usp_DPD_Report '{dpd_table_name.replace(chr(39), chr(39) + chr(39))}'"
                    remember_sql_query(query_text)
                    result, result_meta = run_sql_with_metrics(cursor, query_text)

                    if result:

                        session['last_columns'] = result["columns"]
                        session['last_rows'] = result["rows"]

                    else:

                        conn.commit()

                        result = {
                            "columns": ["Status"],
                            "rows": [["DPD report executed successfully"]]
                        }

                elif action == "run_query":

                    if query_text:

                        remember_sql_query(query_text)
                        result, result_meta = run_sql_with_metrics(cursor, query_text)

                        if result:

                            session['last_columns'] = result["columns"]
                            session['last_rows'] = result["rows"]
                            log_activity("query_run", query_text[:250])

                        else:

                            conn.commit()

                            result = {
                                "columns": ["Status"],
                                "rows": [["Success"]]
                            }

                # ================= ADD TO EXCEL =================

                elif action == "preview_table":

                    if selected_table:
                        query_text = get_table_preview_query(selected_table)
                        result, result_meta = run_sql_with_metrics(cursor, query_text)
                        if result:
                            session['last_columns'] = result["columns"]
                            session['last_rows'] = result["rows"]
                        log_activity("table_previewed", selected_table)

                elif action == "show_columns":

                    if selected_table:
                        started = time.perf_counter()
                        result = get_table_columns_result(cursor, selected_table)
                        result_meta = {
                            "elapsed_ms": int((time.perf_counter() - started) * 1000),
                            "rows_count": len(result["rows"]),
                            "status": "Columns"
                        }
                        session['last_columns'] = result["columns"]
                        session['last_rows'] = result["rows"]
                        log_activity("table_columns_viewed", selected_table)

                elif action == "count_rows":

                    if selected_table:
                        query_text = get_table_count_query(selected_table)
                        result, result_meta = run_sql_with_metrics(cursor, query_text)
                        if result:
                            session['last_columns'] = result["columns"]
                            session['last_rows'] = result["rows"]
                        log_activity("table_rows_counted", selected_table)

                elif action == "validate_sidbi":

                    sidbi_table = dpd_table_name or selected_table
                    if sidbi_table:
                        started = time.perf_counter()
                        result = validate_sidbi_table(cursor, sidbi_table)
                        result_meta = {
                            "elapsed_ms": int((time.perf_counter() - started) * 1000),
                            "rows_count": len(result["rows"]),
                            "status": result["rows"][0][0]
                        }
                        session['last_columns'] = result["columns"]
                        session['last_rows'] = result["rows"]
                        log_activity("sidbi_table_validated", sidbi_table)

                elif action == "compare_query":

                    if query_text and compare_query:
                        left_result, left_meta = run_sql_with_metrics(cursor, query_text)
                        right_result, right_meta = run_sql_with_metrics(cursor, compare_query)
                        left_columns = left_result["columns"] if left_result else []
                        right_columns = right_result["columns"] if right_result else []
                        result = {
                            "columns": ["Metric", "Query 1", "Query 2"],
                            "rows": [
                                ["Rows", left_meta["rows_count"], right_meta["rows_count"]],
                                ["Columns", len(left_columns), len(right_columns)],
                                ["Same Columns", "Yes" if left_columns == right_columns else "No", ""],
                                ["Execution", f'{left_meta["elapsed_ms"]} ms', f'{right_meta["elapsed_ms"]} ms'],
                            ]
                        }
                        result_meta = {
                            "elapsed_ms": left_meta["elapsed_ms"] + right_meta["elapsed_ms"],
                            "rows_count": len(result["rows"]),
                            "status": "Compared"
                        }
                        session['last_columns'] = result["columns"]
                        session['last_rows'] = result["rows"]
                        log_activity("queries_compared", query_text[:120])

                elif action == "add_excel":

                    if multi_query:
                        remember_sql_query(multi_query)

                        queries = [
                            q.strip()
                            for q in multi_query.split(";")
                            if q.strip()
                        ]

                        all_results = session.get(
                            'multi_results',
                            []
                        )

                        for q in queries:

                            try:
                                sheet_name = get_report_sheet_name(
                                    q,
                                    len(all_results) + 1
                                )

                                query_result, query_meta = run_sql_with_metrics(cursor, q)

                                if query_result:
                                    all_results.append({

                                        "sheet_name": sheet_name,

                                        "query": q,
                                        "report_month": report_month,

                                        "columns": query_result["columns"],

                                        "rows": query_result["rows"]
                                    })

                                    result = {
                                        "columns": query_result["columns"],
                                        "rows": query_result["rows"]
                                    }

                                else:

                                    conn.commit()

                                    all_results.append({

                                        "sheet_name": sheet_name,

                                        "query": q,
                                        "report_month": report_month,

                                        "columns": ["Status"],

                                        "rows": [["Success"]]
                                    })

                            except Exception as e:
                                sheet_name = get_report_sheet_name(
                                    q,
                                    len(all_results) + 1
                                )

                                all_results.append({

                                    "sheet_name": sheet_name,

                                    "query": q,
                                    "report_month": report_month,

                                    "columns": ["Error"],

                                    "rows": [[str(e)]]
                                })

                        session['multi_results'] = all_results
                        log_activity("excel_queries_added", f"Added {len(queries)} query results to Excel session")

            except Exception as e:

                error = f"Database error: {e}"

            finally:

                if conn:

                    conn.close()

    return render_template(

        'sql.html',

        result=result,
        error=error,

        servers=servers,
        databases=databases,
        tables=tables,

        selected_server=selected_server,
        selected_database=selected_database,
        selected_table=selected_table,
        selected_uid=selected_uid,
        selected_pwd=selected_pwd,
        query_history=session.get('sql_query_history', []),
        saved_queries=fetch_saved_queries(),
        result_meta=result_meta,
        connection_status=connection_status,

        query=query_text,
        dpd_table_name=dpd_table_name
    )


@app.route('/save_query', methods=['POST'])
@login_required
def save_query():
    title = request.form.get("title", "").strip() or "Saved Query"
    query_text = request.form.get("query_text", "").strip()
    if not query_text:
        flash("Write a query before saving.", "warning")
        return redirect('/sql')

    conn = None
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO dbo.saved_queries (username, title, query_text)
            VALUES (?, ?, ?)
            """,
            (session.get("username"), title, query_text)
        )
        conn.commit()
        log_activity("query_saved", title)
        flash("Query saved successfully.", "success")
    except Exception as e:
        flash(f"Save query failed: {e}", "error")
    finally:
        if conn:
            conn.close()

    return redirect('/sql')


@app.route('/delete_saved_query/<int:query_id>', methods=['POST'])
@login_required
def delete_saved_query(query_id):
    conn = None
    try:
        conn = get_app_conn()
        cursor = conn.cursor()
        cursor.execute(
            "DELETE FROM dbo.saved_queries WHERE id=? AND username=?",
            (query_id, session.get("username"))
        )
        conn.commit()
        log_activity("query_deleted", f"Deleted saved query id: {query_id}")
        flash("Saved query deleted.", "success")
    except Exception as e:
        flash(f"Delete query failed: {e}", "error")
    finally:
        if conn:
            conn.close()

    return redirect('/sql')


# ================= CLEAR =================
@app.route('/clear_multi', methods=['POST'])
@login_required
def clear_multi():
    session.pop("multi_results", None)
    flash("Staged Excel sheets cleared.", "success")
    return redirect('/sql')


# ================= MANAGE =================
@app.route('/manage_sheets')
@login_required
def manage():
    return render_template("manage.html", data=session.get("multi_results", []))


@app.route('/delete_sheet/<int:i>', methods=['POST'])
@login_required
def delete_sheet(i):
    data = session.get("multi_results", [])

    if 0 <= i < len(data):
        data.pop(i)
        flash("Sheet deleted.", "success")
    else:
        flash("Sheet not found.", "error")

    session["multi_results"] = data
    return redirect('/manage_sheets')

@app.route('/view_excel_data')
@login_required
def view_excel_data():

    data = session.get("multi_results", [])

    return render_template("excel_view.html", data=data)    

#-------------------------download_multi_excel----------------------


@app.route('/download_multi_excel')
@login_required
def download_multi_excel():
    data = session.get('multi_results')

    if not data:
        flash("No staged Excel data to download.", "warning")
        return redirect('/sql')

    wb = Workbook()
    wb.remove(wb.active)
    used_sheet_names = set()

    for i, item in enumerate(data):
        sheet_name = unique_excel_sheet_name(
            item.get("sheet_name") or f"Query_{i+1}",
            used_sheet_names
        )
        sheet = wb.create_sheet(title=sheet_name)

        if is_sidbi_top_sheet(item):
            apply_sidbi_top_sheet_format(sheet, item)
        elif is_staff_wise_sidbi_prayas_1_to_30(item):
            apply_staff_wise_sidbi_format(
                sheet,
                item,
                title="STAFF WISE SIDBI PRAYAS 1 TO 30 DAYS"
            )
        elif is_staff_wise_sidbi_prayas(item):
            apply_staff_wise_sidbi_format(sheet, item)
        else:
            sheet.append(item["columns"])

            for row in item["rows"]:
                sheet.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    rows_count = sum(len(item.get("rows", [])) for item in data)
    record_report_history("Final_Multi_Query.xlsx", rows_count, len(data))
    log_activity("excel_downloaded", f"Downloaded multi Excel with {len(data)} sheets")

    return send_file(
        output,
        as_attachment=True,
        download_name="Final_Multi_Query.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



# ------------------ DOWNLOAD EXCEL ------------------
@app.route('/download_excel')
@login_required
def download_excel():

    columns = session.get('last_columns')
    rows = session.get('last_rows')

    if not columns or not rows:
        flash("No query result data to download.", "warning")
        return redirect('/sql')

    wb = Workbook()
    ws = wb.active

    ws.append(columns)

    for row in rows:
        ws.append(row)

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    record_report_history("query_result.xlsx", len(rows), 1)
    log_activity("excel_downloaded", f"Downloaded query_result.xlsx with {len(rows)} rows")

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="query_result.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route('/download_csv')
@login_required
def download_csv():
    columns = session.get('last_columns')
    rows = session.get('last_rows')
    if not columns or rows is None:
        flash("No query result data to download.", "warning")
        return redirect('/sql')
    log_activity("csv_downloaded", f"Downloaded CSV with {len(rows)} rows")
    return export_rows_as_csv(columns, rows)


@app.route('/download_json')
@login_required
def download_json():
    columns = session.get('last_columns')
    rows = session.get('last_rows')
    if not columns or rows is None:
        flash("No query result data to download.", "warning")
        return redirect('/sql')

    data = [dict(zip(columns, row)) for row in rows]
    log_activity("json_downloaded", f"Downloaded JSON with {len(rows)} rows")
    return Response(
        json.dumps(data, default=str, indent=2),
        mimetype="application/json",
        headers={"Content-Disposition": "attachment; filename=query_result.json"}
    )


# ------------------ LOGOUT ------------------
@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')


# ------------------ RUN ------------------
if __name__ == '__main__':
    app.run(debug=True)

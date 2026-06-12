from flask import Flask, render_template as flask_render_template, request, redirect, session, send_file
import pyodbc
import random
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO
from functools import wraps

app = Flask(__name__)
app.secret_key = "secret123"


TEMPLATE_ALIASES = {
    "login.html": "Login.html",
    "add_student.html": "Add_student.html",
    "students.html": "Students.html",
    "settings.html": "Settings.html",
    "error.html": "Error.html",
}


def render_template(template_name, **context):
    template_name = TEMPLATE_ALIASES.get(template_name, template_name)
    context.setdefault("profile_pic", "Default.png")
    context.setdefault(
        "stats",
        {
            "students": 0,
            "users": 0,
            "queries": 0,
            "reports": 0,
            "last_login": None,
            "recent_logs": [],
        },
    )
    context.setdefault("tables", [])
    return flask_render_template("app.html", page_template=template_name, **context)


def login_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect("/login")
        return view_func(*args, **kwargs)

    return wrapper




# ------------------ LOGIN ------------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return render_template('login.html')

    try:
        conn = pyodbc.connect(
            "DRIVER={ODBC Driver 13 for SQL Server};"
            "SERVER=SHUBHAM-monitoring;"
            "DATABASE=testing;"          # ✅ fixed database
            "UID=sa;"
            "PWD=Cashpor@123;"
            "TrustServerCertificate=yes;"
        )
    except Exception as e:
        return f"Database connection failed ❌ {e}"

    cursor = conn.cursor()
    username = request.form['username']
    password = request.form['password']

    cursor.execute(
        "SELECT username FROM users WHERE username=? AND password=?",
        (username, password)
    )
    user = cursor.fetchone()

    if user:
        session['temp_user'] = user[0]
        otp = str(random.randint(100000, 999999))
        session['otp'] = otp
        return "This old App.py file is not configured for SMS OTP. Please run new.py."
        return redirect('/otp')

    return "Invalid Username or Password ❌"

# ------------------ DB CONNECTION ------------------
def get_conn():
    db = session.get('db_config')
    if not db:
        return None

    try:
        conn_str = (
            "DRIVER={ODBC Driver 13 for SQL Server};"
            f"SERVER={db['server']};"
            f"DATABASE={db['database']};"
            f"UID={db['uid']};"
            f"PWD={db['pwd']};"
            "TrustServerCertificate=yes;"
        )
        return pyodbc.connect(conn_str)
    except Exception as e:
        print("DB ERROR:", e)
        return None


# ------------------ HOME ------------------
@app.route('/')
def home():
    return redirect('/login')


# ------------------ DB CONFIG ------------------
@app.route('/db_config')
@login_required
def db_config():
    return render_template('db_config.html')


@app.route('/get_databases', methods=['POST'])
@login_required
def get_databases():
    server = request.form.get('server')
    uid = request.form.get('uid')
    pwd = request.form.get('pwd')

    try:
        conn = pyodbc.connect(
            f"DRIVER={{ODBC Driver 13 for SQL Server}};"
            f"SERVER={server};"
            f"UID={uid};"
            f"PWD={pwd};"
        )

        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sys.databases")

        databases = [row[0] for row in cursor.fetchall()]
        return {"databases": databases}

    except Exception as e:
        return {"error": str(e)}


# ------------------ TABLE LIST ------------------
@app.route('/get_tables')
@login_required
def get_tables():
    conn = get_conn()
    if not conn:
        return {"tables": []}

    cursor = conn.cursor()
    cursor.execute("""
        SELECT TABLE_NAME 
        FROM INFORMATION_SCHEMA.TABLES 
        WHERE TABLE_TYPE='BASE TABLE'
    """)

    tables = [row[0] for row in cursor.fetchall()]
    return {"tables": tables}


# ------------------ CONNECT DB ------------------
@app.route('/connect_db', methods=['POST'])
@login_required
def connect_db():
    session['db_config'] = {
        "server": request.form['server'],
        "database": request.form['database'],
        "uid": request.form['uid'],
        "pwd": request.form['pwd']
    }
    return redirect('/login')
# ------------------ REGISTER PAGE ------------------
@app.route('/register')
def register():
    return render_template('register.html')



# ------------------ REGISTER USER ------------------
@app.route('/register_user', methods=['POST'])
def register_user():

    conn = get_conn()
    if not conn:
        return "Database connection failed ❌"

    try:
        conn = pyodbc.connect(
            "DRIVER={ODBC Driver 13 for SQL Server};"
            "SERVER=SHUBHAM-monitoring;"
            "DATABASE=testing;"          # ✅ fixed database
            "UID=sa;"
            "PWD=Cashpor@123;"
            "TrustServerCertificate=yes;"
        )
        cursor = conn.cursor()

        username = request.form['username'].strip()
        password = request.form['password'].strip()
        email = request.form['email'].strip()
        mobile = request.form['mobile'].strip()

        if not username or not password:
            return "Username & Password required ❌"

        cursor.execute("SELECT * FROM users WHERE username=?", (username,))
        existing = cursor.fetchone()

        if existing:
            return render_template(
                "error.html",
                title="Registration Failed",
                message="User already exists ❌"
            )

        cursor.execute(
            "INSERT INTO users (username, password, email, mobile, profile_pic) VALUES (?, ?, ?, ?, ?)",
            (username, password, email, mobile, 'default.png')
        )

        conn.commit()
        return redirect('/login')

    except Exception as e:
        return f"Error: {str(e)} ❌"

    finally:
        conn.close()


# ------------------ OTP ------------------
@app.route('/otp')
def otp_page():
    if not session.get('otp'):
        return redirect('/login')
    return render_template('otp.html')


@app.route('/verify_otp', methods=['POST'])
def verify_otp():
    if request.form.get('otp') == session.get('otp'):
        session['logged_in'] = True
        session['username'] = session.get('temp_user')
        session.pop('otp', None)
        return redirect('/dashboard')

    return "Wrong OTP ❌"


# ------------------ DASHBOARD ------------------
@app.route('/dashboard')
def dashboard():
    if not session.get('logged_in'):
        return redirect('/login')

    now = datetime.now()

    wish = (
        "Good Morning ☀️" if now.hour < 12 else
        "Good Afternoon 🌤️" if now.hour < 17 else
        "Good Evening 🌙"
    )

    return render_template(
        "dashboard.html",
        username=session.get('username'),
        wish=wish,
        current_date=now.strftime("%d %B %Y")
    )


# ------------------ PROFILE ------------------
@app.route('/profile')
def profile():
    if not session.get('logged_in'):
        return redirect('/login')

    conn = get_conn()
    if not conn:
        return "Database connection failed ❌"

    cursor = conn.cursor()

    cursor.execute(
        "SELECT profile_pic FROM users WHERE username=?",
        (session.get('username'),)
    )

    row = cursor.fetchone()

    profile_pic = row[0] if row and row[0] else "default.png"

    return render_template(
        'profile.html',
        username=session.get('username'),
        profile_pic=profile_pic
    )
@app.route('/add_student')
def add_student():
    if not session.get('logged_in'):
        return redirect('/login')
    return render_template('add_student.html')



# ------------------ ADD STUDENT ------------------
@app.route('/save_student', methods=['POST'])
def save_student():
    print("SAVE STUDENT ROUTE HIT")  # DEBUG

    if not session.get('logged_in'):
        return redirect('/login')

    try:
        # ✅ Fixed connection to Shubham database
        conn = pyodbc.connect(
            "DRIVER={ODBC Driver 13 for SQL Server};"
            "SERVER=SHUBHAM-monitoring;"
            "DATABASE=testing;"          # ✅ fixed database
            "UID=sa;"
            "PWD=Cashpor@123;"
            "TrustServerCertificate=yes;"
        )

        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO dbo.students (name, age, course) VALUES (?, ?, ?)",
            (request.form['name'], request.form['age'], request.form['course'])
        )
        conn.commit()

    except Exception as e:
        return f"Database Error ❌ {e}"

    finally:
        conn.close()

    return redirect('/view_students')


# ------------------ VIEW STUDENTS ------------------
@app.route('/view_students')
def view_students():
    if not session.get('logged_in'):
        return redirect('/login')

    conn = get_conn()
    if not conn:
        return "Database connection failed ❌"

    try:
        # ✅ Fixed connection to Shubham database
        conn = pyodbc.connect(
            "DRIVER={ODBC Driver 13 for SQL Server};"
            "SERVER=SHUBHAM-monitoring;"
            "DATABASE=testing;"          # ✅ fixed database
            "UID=sa;"
            "PWD=Cashpor@123;"
            "TrustServerCertificate=yes;"
        )
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM students")
        students = cursor.fetchall()

    except Exception as e:
        return f"Error: {str(e)}"

    finally:
        conn.close()

    return render_template('students.html', students=students)
# ------------------ DELETE STUDENT ------------------
@app.route('/delete/<int:id>')
def delete_student(id):

    if not session.get('logged_in'):
        return redirect('/login')

    conn = get_conn()
    if not conn:
        return "Database Error ❌"

    try:
        # ✅ Fixed connection to Shubham database
        conn = pyodbc.connect(
            "DRIVER={ODBC Driver 13 for SQL Server};"
            "SERVER=SHUBHAM-monitoring;"
            "DATABASE=testing;"          # ✅ fixed database
            "UID=sa;"
            "PWD=Cashpor@123;"
            "TrustServerCertificate=yes;"
        )
        cursor = conn.cursor()
        cursor.execute("DELETE FROM students WHERE id=?", (id,))
        conn.commit()

    except Exception as e:
        return f"Error: {str(e)}"

    finally:
        conn.close()

    return redirect('/view_students')

# ------------------ DELETE ACCOUNT ------------------
@app.route('/delete_account', methods=['POST'])
def delete_account():
    if not session.get('logged_in'):
        return redirect('/login')

    conn = get_conn()
    if not conn:
        return "DB error ❌"

    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM users WHERE username=?",
        (session.get('username'),)
    )
    conn.commit()

    session.clear()
    return redirect('/login')


# ------------------ SETTINGS ------------------
@app.route('/settings')
def settings():
    if not session.get('logged_in'):
        return redirect('/login')
    return render_template('settings.html')


# ------------------ UPDATE PROFILE ------------------
@app.route('/update_profile', methods=['POST'])
def update_profile():
    if not session.get('logged_in'):
        return redirect('/login')

    conn = get_conn()
    if not conn:
        return "Database connection failed ❌"

    new_username = request.form['username']
    new_pic = request.form['profile_pic']

    cursor = conn.cursor()
    cursor.execute(
        "UPDATE users SET username=?, profile_pic=? WHERE username=?",
        (new_username, new_pic, session.get('username'))
    )
    conn.commit()

    session['username'] = new_username

    return redirect('/profile')


# ------------------ CHANGE PASSWORD ------------------
@app.route('/change_password', methods=['POST'])
def change_password():
    if not session.get('logged_in'):
        return redirect('/login')

    conn = get_conn()
    if not conn:
        return "DB error ❌"

    old_pass = request.form['old_password']
    new_pass = request.form['new_password']

    cursor = conn.cursor()

    cursor.execute(
        "SELECT * FROM users WHERE username=? AND password=?",
        (session.get('username'), old_pass)
    )

    user = cursor.fetchone()

    if user:
        cursor.execute(
            "UPDATE users SET password=? WHERE username=?",
            (new_pass, session.get('username'))
        )
        conn.commit()
        return render_template("error.html",
                       title="Password",
                       message="Password Changed Successfully ✅")
       

    return "Wrong Old Password ❌"


# ------------------ SQL RUNNER -----------------

@app.route('/sql', methods=['GET', 'POST'])
def sql_runner():
    if not session.get('logged_in'):
        return redirect('/login')

    conn = get_conn()
    if not conn:
        return "❌ Database not connected"

    cursor = conn.cursor()

    result = None
    error = None
    query_text = ""

    # 🔥 TABLE FETCH (SAFE + ORDERED)
    try:
        cursor.execute("""
            SELECT TABLE_NAME 
            FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_TYPE = 'BASE TABLE'
            ORDER BY TABLE_NAME
        """)
        tables = [row[0] for row in cursor.fetchall()]
    except Exception as e:
        tables = []
        print("Table Fetch Error:", e)

    if request.method == 'POST':

        query_text = request.form.get('query', "").strip()
        multi_query = request.form.get('multi_query', "").strip()

        # =====================================
        # 🔹 SINGLE QUERY
        # =====================================
        if query_text:
            try:
                cursor.execute(query_text)

                if cursor.description:
                    columns = [col[0] for col in cursor.description]
                    rows = [list(r) for r in cursor.fetchall()]
                    result = {"columns": columns, "rows": rows}
                else:
                    conn.commit()
                    result = {"columns": ["Status"], "rows": [["Success"]]}

            except Exception as e:
                error = str(e)

        # =====================================
        # 🔥 MULTI QUERY (EXCEL BUILD MODE)
        # =====================================
        if multi_query:
            queries = [q.strip() for q in multi_query.split(";") if q.strip()]

            all_results = session.get('multi_results', [])

            for i, q in enumerate(queries, start=1):
                try:
                    cursor.execute(q)

                    if cursor.description:
                        columns = [col[0] for col in cursor.description]
                        rows = [list(r) for r in cursor.fetchall()]

                        all_results.append({
                            "sheet_name": f"Query_{len(all_results)+1}",
                            "query": q,
                            "columns": columns,
                            "rows": rows
                        })

                        # last result show karo
                        result = {"columns": columns, "rows": rows}

                    else:
                        conn.commit()
                        all_results.append({
                            "sheet_name": f"Query_{len(all_results)+1}",
                            "query": q,
                            "columns": ["Status"],
                            "rows": [["Success"]]
                        })

                except Exception as e:
                    all_results.append({
                        "sheet_name": f"Query_{len(all_results)+1}",
                        "query": q,
                        "columns": ["Error"],
                        "rows": [[str(e)]]
                    })

            # 🔥 SAVE IN SESSION
            session['multi_results'] = all_results

    return render_template(
        "sql.html",
        result=result,
        error=error,
        tables=tables,
        query=query_text   # 🔥 important for textarea retain
    )


# ================= CLEAR =================
@app.route('/clear_multi')
def clear_multi():
    session.pop("multi_results", None)
    return redirect('/sql')


# ================= MANAGE =================
@app.route('/manage_sheets')
def manage():
    return render_template("manage.html", data=session.get("multi_results", []))


@app.route('/delete_sheet/<int:i>')
def delete_sheet(i):
    data = session.get("multi_results", [])

    if 0 <= i < len(data):
        data.pop(i)

    session["multi_results"] = data
    return redirect('/manage_sheets')

@app.route('/view_excel_data')
def view_excel_data():

    data = session.get("multi_results", [])

    return render_template("excel_view.html", data=data)    

#-------------------------download_multi_excel----------------------


@app.route('/download_multi_excel')
def download_multi_excel():
    from openpyxl import Workbook
    from io import BytesIO

    data = session.get('multi_results')

    if not data:
        return "❌ No data to download"

    wb = Workbook()
    wb.remove(wb.active)

    for i, item in enumerate(data):
        sheet = wb.create_sheet(title=f"Query_{i+1}")

        sheet.append(item["columns"])

        for row in item["rows"]:
            sheet.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="Final_Multi_Query.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



# ------------------ DOWNLOAD EXCEL ------------------
@app.route('/download_excel')
def download_excel():

    columns = session.get('last_columns')
    rows = session.get('last_rows')

    if not columns or not rows:
        return "No data to download ❌"

    wb = Workbook()
    ws = wb.active

    ws.append(columns)

    for row in rows:
        ws.append(row)

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="query_result.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ------------------ LOGOUT ------------------
@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')


# ------------------ RUN ------------------
if __name__ == '__main__':
    app.run(debug=True)
